from flask import Flask, render_template, request, redirect, session, Response
import random
import sqlite3
import datetime
import ast
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import io
from fragen import fragen

app = Flask(__name__)
app.secret_key = "supersecretkey"

ADMIN_PASSWORD = "IKLMZ"


# ---------------------------
# DB
# ---------------------------
def init_db():
    conn = sqlite3.connect("daten.db")
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS results (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            buddy TEXT,
            time TEXT,
            answers TEXT,
            result TEXT
        )
    """)
    conn.commit()
    conn.close()

init_db()

# ---------------------------
# START
# ---------------------------
@app.route("/", methods=["GET", "POST"])
def start():
    if request.method == "POST":
        session["name"] = request.form["name"]
        session["buddy"] = request.form.get("buddy")
        session["index"] = 0
        session["answers"] = []
        session["saved"] = False
        return redirect("/frage")
    return render_template("start.html")

# ---------------------------
# FRAGEN
# ---------------------------
@app.route("/frage", methods=["GET", "POST"])
def frage():
    if "answers" not in session:
        return redirect("/")

    i = session.get("index", 0)

    if i >= len(fragen):
        return redirect("/ende")

    frage = fragen[i]
    antworten = frage["antworten"].copy()
    random.shuffle(antworten)

    if request.method == "POST":
        selected = request.form.getlist("antwort")

        if not selected:
            return render_template(
                "frage.html",
                frage=frage,
                antworten=antworten,
                error="Bitte wähle eine Antwort"
            )

        # 🔥 WICHTIG: wir speichern TEXT + KATEGORIE sauber getrennt
        chosen = []

        for a in frage["antworten"]:
            if a["kategorie"] in selected:
                chosen.append({
                    "god": a["kategorie"],
                    "text": a["text"]
                })

        session["answers"].append({
            "frage": frage["frage"],
            "selected": chosen
        })

        session["index"] = i + 1
        return redirect("/frage")

    return render_template("frage.html", frage=frage, antworten=antworten)

# ---------------------------
# AUSWERTUNG
# ---------------------------
def berechne_result(answers):
    counter = {
        "Hades": 0,
        "Poseidon": 0,
        "Demeter": 0,
        "Athene": 0,
        "Apollo": 0
    }

    for eintrag in answers:
        antworten = eintrag.get("selected", [])

        for a in antworten:
            god = a.get("god")
            if god:
                counter[god] += 1

    return {k: v * 10 for k, v in counter.items()}

# ---------------------------
# ENDE
# ---------------------------
@app.route("/ende")
def ende():
    # 🔥 Schon gespeichert?
    if session.get("saved"):
        return render_template("ende.html")

    name = session.get("name")
    buddy = session.get("buddy")
    answers = session.get("answers")

    result = berechne_result(answers)

    conn = sqlite3.connect("daten.db")
    c = conn.cursor()

    c.execute(
        "INSERT INTO results (name, buddy, time, answers, result) VALUES (?, ?, ?, ?, ?)",
        (
            name,
            buddy,
            (datetime.datetime.now() + datetime.timedelta(hours=2)).strftime("%d.%m.%Y %H:%M"),
            str(answers),
            str(result)
        )
    )

    conn.commit()
    conn.close()

    # 🔥 markieren als gespeichert
    session["saved"] = True

    return render_template("ende.html")

# ---------------------------
# ADMIN LOGIN
# ---------------------------
@app.route("/admin", methods=["GET", "POST"])
def admin():
    if request.method == "POST":
        if request.form["password"] == ADMIN_PASSWORD:
            session["admin"] = True
            return redirect("/admin/dashboard")
        return "Falsches Passwort"

    return render_template("admin_login.html")

# ---------------------------
# ADMIN DASHBOARD
# ---------------------------
@app.route("/admin/dashboard")
def dashboard():
    if not session.get("admin"):
        return redirect("/admin")

    conn = sqlite3.connect("daten.db")
    c = conn.cursor()

    c.execute("SELECT name, buddy, time, result FROM results")
    rows = c.fetchall()
    conn.close()

    data = []
    for r in rows:
        try:
            result = ast.literal_eval(r[3])
        except:
            result = {}

        data.append((r[0], r[1], r[2], result))

    return render_template("admin.html", data=data)

# ---------------------------
# ADMIN RESET
# ---------------------------
@app.route("/admin/reset", methods=["POST"])
def reset():
    if not session.get("admin"):
        return redirect("/admin")

    conn = sqlite3.connect("daten.db")
    c = conn.cursor()
    c.execute("DELETE FROM results")
    conn.commit()
    conn.close()

    return redirect("/admin/dashboard")

# ---------------------------
# ADMIN CARDS
# ---------------------------
@app.route("/admin/cards")
def cards():
    if not session.get("admin"):
        return redirect("/admin")

    conn = sqlite3.connect("daten.db")
    c = conn.cursor()
    c.execute("SELECT name, buddy, time, result, answers FROM results")
    rows = c.fetchall()
    conn.close()

    import ast

    data = []
    for r in rows:
        name = r[0]
        buddy = r[1]
        time = r[2]
        result = ast.literal_eval(r[3])
        answers = ast.literal_eval(r[4])

        data.append((name, buddy, time, result, answers))

    return render_template("admin_cards.html", data=data)
    

# ---------------------------
# EXCEL
# ---------------------------

@app.route("/admin/export")
def export_excel():

    if not session.get("admin"):
        return redirect("/admin")

    conn = sqlite3.connect("daten.db")
    c = conn.cursor()

    c.execute(
        "SELECT name, buddy, result FROM results"
    )

    rows = c.fetchall()

    conn.close()

    # =========================================================
    # EXCEL WORKBOOK
    # =========================================================

    wb = Workbook()


    # =========================================================
    # SHEET 1: ERGEBNISSE
    # =========================================================

    ws = wb.active
    ws.title = "Ergebnisse"

    gods = [
        "Athene",
        "Poseidon",
        "Demeter",
        "Hades",
        "Apollo"
    ]

    # Header

    ws.append(
        ["Name"] + gods + ["Buddy"]
    )

    green_fill = PatternFill(
        start_color="C6EFCE",
        end_color="C6EFCE",
        fill_type="solid"
    )

    # Personen eintragen

    for r in rows:

        name = r[0]
        buddy = r[1]

        result = ast.literal_eval(r[2])

        row_values = [name]

        max_value = (
            max(result.values())
            if result
            else 0
        )

        # Gott-Werte

        for g in gods:

            row_values.append(
                result.get(g, 0)
            )

        row_values.append(buddy)

        ws.append(row_values)

        # Höchsten Wert grün markieren

        row_index = ws.max_row

        for col_index, g in enumerate(
            gods,
            start=2
        ):

            cell = ws.cell(
                row=row_index,
                column=col_index
            )

            if (
                cell.value == max_value
                and max_value > 0
            ):

                cell.fill = green_fill


    # =========================================================
    # ZUTEILUNG VORBEREITEN
    # =========================================================

    personen = []

    for r in rows:

        personen.append({
            "name": r[0],
            "buddy": r[1],
            "result": ast.literal_eval(r[2])
        })


    # ---------------------------------------------------------
    # Namen vereinheitlichen
    # Groß-/Kleinschreibung egal
    # ---------------------------------------------------------

    def normalisiere(name):

        if not name:
            return ""

        return str(name).strip().lower()


    personen_namen = {
        normalisiere(p["name"]): p
        for p in personen
    }


    # =========================================================
    # TANDEMS FINDEN
    # =========================================================

    gruppen = []

    bereits_verwendet = set()

    for person in personen:

        name_key = normalisiere(
            person["name"]
        )

        if name_key in bereits_verwendet:
            continue

        buddy_name = normalisiere(
            person["buddy"]
        )

        # Gibt es den Buddy als Teilnehmer?

        if (
            buddy_name
            and buddy_name in personen_namen
        ):

            buddy_person = personen_namen[
                buddy_name
            ]

            buddy_key = normalisiere(
                buddy_person["name"]
            )

            if buddy_key != name_key:

                gruppen.append({
                    "personen": [
                        person,
                        buddy_person
                    ],
                    "tandem": True
                })

                bereits_verwendet.add(
                    name_key
                )

                bereits_verwendet.add(
                    buddy_key
                )

                continue


        # Einzelperson

        gruppen.append({
            "personen": [person],
            "tandem": False
        })

        bereits_verwendet.add(
            name_key
        )


    # =========================================================
    # TEAMGRÖSSEN
    # =========================================================

    anzahl_personen = len(personen)

    basisgroesse = (
        anzahl_personen // len(gods)
    )

    rest = (
        anzahl_personen % len(gods)
    )

    kapazitaeten = {}

    for i, god in enumerate(gods):

        kapazitaeten[god] = (
            basisgroesse
            + (
                1
                if i < rest
                else 0
            )
        )


    # =========================================================
    # HILFSFUNKTION FÜR TEAM-SCORE
    # =========================================================

    def gruppen_score(
        gruppe,
        god
    ):

        if not gruppe["personen"]:
            return 0

        gesamt = 0

        for person in gruppe["personen"]:

            gesamt += person[
                "result"
            ].get(god, 0)

        # Durchschnitt
        return (
            gesamt
            / len(gruppe["personen"])
        )


    # =========================================================
    # TEAMS
    # =========================================================

    teams = {
        god: []
        for god in gods
    }

    team_anzahl = {
        god: 0
        for god in gods
    }


    # =========================================================
    # TANDEMS ZUERST
    # =========================================================

    tandems = [
        g
        for g in gruppen
        if g["tandem"]
    ]

    einzelpersonen = [
        g
        for g in gruppen
        if not g["tandem"]
    ]


    # Stärkste Tandems zuerst

    tandems.sort(
        key=lambda g: max(
            gruppen_score(g, god)
            for god in gods
        ),
        reverse=True
    )


    for gruppe in tandems:

        moegliche_teams = []

        for god in gods:

            groesse = len(
                gruppe["personen"]
            )

            if (
                team_anzahl[god]
                + groesse
                <= kapazitaeten[god]
            ):

                score = gruppen_score(
                    gruppe,
                    god
                )

                moegliche_teams.append(
                    (
                        score,
                        god
                    )
                )


        if moegliche_teams:

            moegliche_teams.sort(
                key=lambda x: x[0],
                reverse=True
            )

            bestes_team = (
                moegliche_teams[0][1]
            )

        else:

            bestes_team = min(
                gods,
                key=lambda g:
                    team_anzahl[g]
            )


        teams[
            bestes_team
        ].extend(
            gruppe["personen"]
        )

        team_anzahl[
            bestes_team
        ] += len(
            gruppe["personen"]
        )


    # =========================================================
    # EINZELPERSONEN
    # =========================================================

    einzelpersonen.sort(
        key=lambda g: max(
            g["personen"][0][
                "result"
            ].get(god, 0)
            for god in gods
        ),
        reverse=True
    )


    for gruppe in einzelpersonen:

        person = gruppe[
            "personen"
        ][0]

        moegliche_teams = []

        for god in gods:

            if (
                team_anzahl[god]
                < kapazitaeten[god]
            ):

                score = person[
                    "result"
                ].get(god, 0)

                moegliche_teams.append(
                    (
                        score,
                        god
                    )
                )


        if moegliche_teams:

            moegliche_teams.sort(
                key=lambda x: x[0],
                reverse=True
            )

            bester_score = (
                moegliche_teams[0][0]
            )

            # Teams, die höchstens 10 %
            # schlechter sind, gelten
            # als ähnlich gut.

            gleich_gute = [
                god
                for score, god
                in moegliche_teams
                if score
                >= bester_score - 10
            ]

            # Bei ähnlichem Ergebnis:
            # kleineres Team bevorzugen.

            bestes_team = min(
                gleich_gute,
                key=lambda g:
                    team_anzahl[g]
            )

        else:

            bestes_team = min(
                gods,
                key=lambda g:
                    team_anzahl[g]
            )


        teams[
            bestes_team
        ].append(person)

        team_anzahl[
            bestes_team
        ] += 1


    # =========================================================
    # SHEET 2: ZUTEILUNG
    # =========================================================

    ws2 = wb.create_sheet(
        "Zuteilung"
    )


    # ---------------------------------------------------------
    # Kleine Übersicht oben
    # ---------------------------------------------------------

    ws2["A1"] = "Zuteilung"

    ws2["A1"].font = (
        ws2["A1"].font.copy(
            bold=True,
            size=16
        )
    )

    ws2["A3"] = "Team"
    ws2["B3"] = "Anzahl"


    for row_index, god in enumerate(
        gods,
        start=4
    ):

        ws2.cell(
            row=row_index,
            column=1,
            value=god
        )

        ws2.cell(
            row=row_index,
            column=2,
            value=len(
                teams[god]
            )
        )


    # ---------------------------------------------------------
    # Eigentliche Tabelle
    # ---------------------------------------------------------

    start_row = 11

    headers = (
        ["Name"]
        + gods
        + ["Buddy"]
    )


    # Tabellenkopf

    for col_index, header in enumerate(
        headers,
        start=1
    ):

        cell = ws2.cell(
            row=start_row,
            column=col_index,
            value=header
        )

        cell.font = (
            cell.font.copy(
                bold=True
            )
        )


    start_row += 1


    # =========================================================
    # TEAMS UNTEREINANDER
    # =========================================================

    for god in gods:

        # Teamüberschrift

        team_cell = ws2.cell(
            row=start_row,
            column=1,
            value=god
        )

        team_cell.font = (
            team_cell.font.copy(
                bold=True,
                size=14
            )
        )

        start_row += 1


        # Personen dieses Teams

        for person in teams[god]:

            name = person["name"]
            buddy = person["buddy"]
            result = person["result"]


            # Name

            ws2.cell(
                row=start_row,
                column=1,
                value=name
            )


            # Prozentwerte

            for col_index, g in enumerate(
                gods,
                start=2
            ):

                ws2.cell(
                    row=start_row,
                    column=col_index,
                    value=result.get(
                        g,
                        0
                    )
                )


            # Buddy

            ws2.cell(
                row=start_row,
                column=7,
                value=buddy
            )


            # Höchsten Wert grün markieren

            max_value = (
                max(
                    result.values()
                )
                if result
                else 0
            )


            for col_index, g in enumerate(
                gods,
                start=2
            ):

                cell = ws2.cell(
                    row=start_row,
                    column=col_index
                )

                if (
                    cell.value
                    == max_value
                    and max_value > 0
                ):

                    cell.fill = green_fill


            start_row += 1


        # Abstand zwischen Teams

        start_row += 2


    # =========================================================
    # SPALTENBREITEN
    # =========================================================

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 14
    ws2.column_dimensions["F"].width = 14
    ws2.column_dimensions["G"].width = 25


    # =========================================================
    # DATEI AUSGEBEN
    # =========================================================

    output = io.BytesIO()

    wb.save(output)

    output.seek(0)

    return Response(
        output,
        mimetype=(
            "application/vnd.openxmlformats-"
            "officedocument.spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition":
                "attachment;"
                "filename=ergebnisse.xlsx"
        }
    )



# ---------------------------
# ADMIN FRAGEN AUSWERTUNG
# ---------------------------
@app.route("/admin/fragen")
def admin_fragen():
    if not session.get("admin"):
        return redirect("/admin")

    conn = sqlite3.connect("daten.db")
    c = conn.cursor()
    c.execute("SELECT answers FROM results")
    rows = c.fetchall()
    conn.close()

    # 🔢 Zähler vorbereiten
    stats = {}

    for f in fragen:
        stats[f["frage"]] = {}
        for a in f["antworten"]:
            stats[f["frage"]][a["text"]] = {
                "god": a["kategorie"],
                "count": 0
            }

    total_people = len(rows)

    # 🔢 Antworten zählen
    for r in rows:
        try:
            answers = ast.literal_eval(r[0])
        except:
            continue

        for entry in answers:
            frage_text = entry["frage"]

            for s in entry["selected"]:
                answer_text = s["text"]

                if frage_text in stats and answer_text in stats[frage_text]:
                    stats[frage_text][answer_text]["count"] += 1

    # 📊 Prozent berechnen
    for frage_text in stats:
        for answer_text in stats[frage_text]:
            count = stats[frage_text][answer_text]["count"]

            if total_people > 0:
                percent = round(count / total_people * 100, 1)
            else:
                percent = 0

            stats[frage_text][answer_text]["percent"] = percent

    return render_template("admin_fragen.html", stats=stats)

# ---------------------------
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
